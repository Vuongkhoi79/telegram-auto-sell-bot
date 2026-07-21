"""Import account inventory from an .xlsx or .csv file into database/store.db.

This script is not wired into the Telegram bot.  Run with --help for usage.
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sqlite3
import sys
import tempfile
import uuid
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from repository.store_repository import ACCOUNT_PRODUCT_CODE_ALIASES, canonical_catalog_product


DEFAULT_DATABASE = PROJECT_ROOT / "database" / "store.db"
REQUIRED_COLUMNS = (
    "product_code", "category", "product_name", "account_type", "duration",
    "price_vnd", "warranty_days", "credential_text", "note", "active",
)
OPTIONAL_COLUMNS = (
    "category_key", "menu_order", "show_in_menu", "product_group", "description",
    "package_name", "plan_name", "email", "password", "2fa", "recovery_email",
)
IMPORT_COLUMNS = tuple(dict.fromkeys((*REQUIRED_COLUMNS, *OPTIONAL_COLUMNS)))
SKIPPED_SHEET_NAMES = {"README", "HƯỚNG DẪN", "HUONG DAN", "TEMPLATE"}
DEFAULT_PRODUCT_DISPLAY_NAMES = {
    "CAPCUT": "CAPCUT PRO 12M",
    "CAPCUT_7D": "CAPCUT PRO 7 ngay",
    "CAPCUT_12M": "CAPCUT PRO 12M",
    "CAPCUT_30D": "CAPCUT PRO 30D",
    "GEMINI": "Gemini AI Pro",
    "CHATGPT_SHARED": "ChatGPT Plus dùng chung",
}
EXPECTED_PRODUCT_TERMS = {
    "CAPCUT": {"price_vnd": 400000, "warranty_days": 365},
    "CAPCUT_7D": {"price_vnd": 8000, "warranty_days": 7},
    "CAPCUT_12M": {"price_vnd": 400000, "warranty_days": 365},
    "CAPCUT_30D": {"price_vnd": 45000, "warranty_days": 30},
    "CHATGPT_SHARED": {"price_vnd": 45000, "warranty_days": 7},
}
ALLOWED_PRODUCT_CODES = {
    "CHATGPT",
    "CHATGPT_SHARED",
    "GEMINI",
    "GROK",
    "CAPCUT",
    "CAPCUT_7D",
    "CAPCUT_12M",
    "CAPCUT_30D",
    "CLAUDE",
    "CURSOR",
    "CANVA",
    "ADOBE",
    "ARTLIST",
    "ELEVEN",
    "GAMMA",
    "HEYGEN",
    "HIGGSFIELD",
    "KLING",
    "KREA",
    "OPENART",
    "SUNO",
    "VEO3",
    "VIEWMAX",
}


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def validate_credential(value: Any) -> str:
    credential = text(value)
    parts = credential.split("|")
    if len(parts) not in {2, 3, 4} or any(not part.strip() for part in parts):
        raise ValueError("credential_text must be email|password, email|password|2fa, or email|password|2fa|recovery_email")
    return credential


def credential_from_row(row: dict[str, Any]) -> str:
    credential_text = text(row.get("credential_text"))
    if credential_text:
        return validate_credential(credential_text)

    email = text(row.get("email"))
    password = text(row.get("password"))
    two_factor = text(row.get("2fa"))
    recovery_email = text(row.get("recovery_email"))
    if not email or not password:
        raise ValueError("credential_text or email/password is required")
    parts = [email, password]
    if two_factor or recovery_email:
        parts.append(two_factor)
    if recovery_email:
        parts.append(recovery_email)
    return validate_credential("|".join(parts))


def validate_product_code(product_code: str) -> None:
    if product_code not in ALLOWED_PRODUCT_CODES:
        allowed = ", ".join(sorted(ALLOWED_PRODUCT_CODES))
        raise ValueError(f"Unsupported product_code: {product_code}. Allowed product_code values: {allowed}")


def validate_product_terms(product_code: str, row: dict[str, Any]) -> None:
    expected = EXPECTED_PRODUCT_TERMS.get(product_code)
    if not expected:
        return
    price_vnd = parse_int(row.get("price_vnd"), 0)
    warranty_days = parse_int(row.get("warranty_days"), 0)
    if price_vnd != expected["price_vnd"] or warranty_days != expected["warranty_days"]:
        raise ValueError(
            f"{product_code} requires price_vnd={expected['price_vnd']} "
            f"and warranty_days={expected['warranty_days']}; got price_vnd={price_vnd}, "
            f"warranty_days={warranty_days}"
        )


def normalize_import_product_code(product_code: str, row: dict[str, Any]) -> str:
    normalized = text(product_code).upper()
    duration = text(row.get("duration")).upper().replace(" ", "")
    price_vnd = parse_int(row.get("price_vnd"), 0)
    warranty_days = parse_int(row.get("warranty_days"), 0)
    if normalized == "CAPCUT" and duration in {"7D", "7DAY", "7DAYS", "7NGAY", "7NGÀY"} and price_vnd == 8000 and warranty_days == 7:
        return "CAPCUT_7D"
    if normalized == "CAPCUT" and duration in {"30D", "30DAY", "30DAYS", "30NGAY", "30NGÃ€Y"} and price_vnd == 45000 and warranty_days == 30:
        return "CAPCUT_30D"
    return normalized


def canonical_product_code(product_code: str) -> str:
    normalized = text(product_code).upper()
    if normalized in ALLOWED_PRODUCT_CODES:
        return normalized
    for canonical, aliases in ACCOUNT_PRODUCT_CODE_ALIASES.items():
        if normalized in {alias.upper() for alias in aliases}:
            return canonical
    return normalized


def product_code_candidates(product_code: str) -> list[str]:
    if product_code in {"CAPCUT_12M", "CAPCUT_30D", "CAPCUT_7D"}:
        return [product_code]
    candidates = [product_code]
    for alias in ACCOUNT_PRODUCT_CODE_ALIASES.get(product_code, ()):
        alias = alias.upper()
        if alias not in candidates:
            candidates.append(alias)
    return candidates


def find_existing_product(connection: sqlite3.Connection, product_code: str) -> sqlite3.Row | None:
    for candidate in product_code_candidates(product_code):
        row = connection.execute(
            "SELECT id, code FROM products WHERE code = ? AND active = 1",
            (candidate,),
        ).fetchone()
        if row:
            return row
    return None


def find_product_by_code(connection: sqlite3.Connection, product_code: str) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT id, code FROM products WHERE code = ? AND active = 1",
        (product_code,),
    ).fetchone()


def product_display_name(product_code: str, row: dict[str, Any]) -> str:
    canonical = canonical_catalog_product(product_code)
    if canonical:
        return text(canonical["name"])
    return (
        text(row.get("package_name"))
        or text(row.get("plan_name"))
        or DEFAULT_PRODUCT_DISPLAY_NAMES.get(product_code)
        or text(row.get("product_name"))
        or product_code
    )


def canonical_price_vnd(product_code: str, row: dict[str, Any]) -> int:
    canonical = canonical_catalog_product(product_code)
    if canonical:
        return int(canonical["price_vnd"])
    return parse_int(row.get("price_vnd"), 0)


def parse_int(value: Any, default: int = 0) -> int:
    raw = text(value)
    if not raw:
        return default
    return int(float(raw))


def sync_product_metadata(
    connection: sqlite3.Connection,
    product_id: str,
    product_code: str,
    row: dict[str, Any],
    now: str,
) -> None:
    connection.execute(
        """
        UPDATE products
        SET name = ?, updated_at = ?, category = 'account', account_type = ?,
            duration = ?, price_vnd = ?, warranty_days = ?, note = ?,
            category_key = ?, product_group = 'account', active = 1, delivery_type = 'account'
        WHERE id = ?
        """,
        (
            product_display_name(product_code, row),
            now,
            text(row.get("account_type")),
            text(row.get("duration")),
            canonical_price_vnd(product_code, row),
            parse_int(row.get("warranty_days"), 0),
            text(row.get("note")),
            product_code,
            product_id,
        ),
    )


def ensure_product(
    connection: sqlite3.Connection,
    product_code: str,
    row: dict[str, Any],
    now: str,
) -> sqlite3.Row:
    product = find_product_by_code(connection, product_code)
    if product:
        sync_product_metadata(connection, product["id"], product_code, row, now)
        return product

    product_id = str(uuid.uuid4())
    connection.execute(
        """
        INSERT INTO products
            (id, code, name, active, delivery_type, created_at, updated_at,
             category, account_type, duration, price_vnd, warranty_days, note,
             category_key, product_group)
        VALUES (?, ?, ?, 1, 'account', ?, ?, 'account', ?, ?, ?, ?, ?, ?, 'account')
        """,
        (
            product_id,
            product_code,
            product_display_name(product_code, row),
            now,
            now,
            text(row.get("account_type")),
            text(row.get("duration")),
            canonical_price_vnd(product_code, row),
            parse_int(row.get("warranty_days"), 0),
            text(row.get("note")),
            product_code,
        ),
    )
    return connection.execute(
        "SELECT id, code FROM products WHERE id = ?",
        (product_id,),
    ).fetchone()


def read_rows(path: Path) -> Iterator[tuple[int, dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            if not reader.fieldnames:
                raise ValueError("CSV has no header row")
            fields = {field.strip().lower() for field in reader.fieldnames if field}
            if "product_code" not in fields:
                raise ValueError("Missing required column: product_code")
            if "credential_text" not in fields and not {"email", "password"}.issubset(fields):
                raise ValueError("Missing credential columns: use credential_text or email/password")
            for row_number, row in enumerate(reader, start=2):
                normalized_row = {str(key).strip().lower(): value for key, value in row.items() if key}
                yield row_number, {key: normalized_row.get(key, "") for key in IMPORT_COLUMNS}
        return
    if suffix != ".xlsx":
        raise ValueError("Input must be an .xlsx or .csv file")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            sheet_code = text(sheet.title).upper()
            if sheet_code in SKIPPED_SHEET_NAMES:
                continue
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            header = [text(value).lower() for value in header_row]
            positions = {name: index for index, name in enumerate(header) if name}
            if "credential_text" not in positions and not {"email", "password"}.issubset(positions):
                raise ValueError(f"Sheet {sheet.title}: missing credential columns: use credential_text or email/password")
            sheet_is_product = sheet_code in ALLOWED_PRODUCT_CODES
            if not sheet_is_product and "product_code" not in positions:
                raise ValueError(f"Sheet {sheet.title}: missing required column: product_code")
            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value is not None and text(value) for value in values):
                    continue
                row = {
                    column: values[positions[column]] if column in positions and positions[column] < len(values) else ""
                    for column in IMPORT_COLUMNS
                }
                if sheet_is_product:
                    row_product_code = text(row.get("product_code")).upper()
                    if not row_product_code or row_product_code == sheet_code:
                        row["product_code"] = sheet_code
                    row["__sheet_product"] = True
                yield row_number, row
    finally:
        workbook.close()


def disable_available_inventory(connection: sqlite3.Connection, product_id: str, now: str) -> int:
    rows = connection.execute(
        "SELECT id FROM inventory_items WHERE product_id = ? AND status = 'available'",
        (product_id,),
    ).fetchall()
    for row in rows:
        connection.execute(
            "UPDATE inventory_items SET status = 'disabled', disabled_at = ? WHERE id = ?",
            (now, row["id"]),
        )
        connection.execute(
            """
            INSERT INTO inventory_movements
                (id, inventory_item_id, action, source, created_at)
            VALUES (?, ?, 'disable', 'import_inventory_replace', ?)
            """,
            (str(uuid.uuid4()), row["id"], now),
        )
    return len(rows)


def active_product_ids_for_code(connection: sqlite3.Connection, product_code: str) -> list[str]:
    product_ids: list[str] = []
    for candidate in product_code_candidates(product_code):
        row = find_product_by_code(connection, candidate)
        if row and row["id"] not in product_ids:
            product_ids.append(row["id"])
    return product_ids


def available_stock_for_code(connection: sqlite3.Connection, product_code: str) -> int:
    exact = find_product_by_code(connection, product_code)
    if exact:
        row = connection.execute(
            "SELECT COUNT(*) AS stock FROM inventory_items WHERE product_id = ? AND status = 'available'",
            (exact["id"],),
        ).fetchone()
        return int(row["stock"] or 0)

    total = 0
    for candidate in product_code_candidates(product_code):
        row = connection.execute(
            """
            SELECT COUNT(*) AS stock
            FROM inventory_items AS i
            JOIN products AS p ON p.id = i.product_id
            WHERE p.code = ? AND i.status = 'available'
            """,
            (candidate,),
        ).fetchone()
        total += int(row["stock"] or 0)
    return total


def should_use_canonical_product(product_code: str, row: dict[str, Any]) -> bool:
    return product_code.startswith("CAPCUT") or product_code == "CHATGPT_SHARED" or bool(row.get("__sheet_product"))


def import_inventory(input_path: Path, database_path: Path = DEFAULT_DATABASE, mode: str = "replace") -> dict[str, Any]:
    if mode not in {"append", "replace"}:
        raise ValueError("mode must be append or replace")
    if not database_path.is_file():
        raise FileNotFoundError(f"Store database not found: {database_path}")
    report: dict[str, Any] = {
        "products_created": 0, "products_updated": 0, "credentials_added": 0,
        "credentials_duplicate": 0, "credentials_disabled": 0, "row_errors": 0, "invalid_rows": 0,
        "rows_read": 0, "valid_rows": 0, "errors": [], "stock": {},
    }
    rows = list(read_rows(input_path))
    report["rows_read"] = len(rows)
    product_seed_rows: dict[str, dict[str, Any]] = {}
    for _, row in rows:
        product_code = normalize_import_product_code(text(row.get("product_code")), row)
        row["product_code"] = product_code
        if product_code and product_code not in product_seed_rows:
            product_seed_rows[product_code] = row

    with closing(sqlite3.connect(database_path)) as connection, connection:
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        replaced_product_ids: set[str] = set()
        imported_codes: list[str] = []
        imported_code_seen: set[str] = set()

        def add_imported_code(product_code: str) -> None:
            canonical = canonical_product_code(product_code)
            if canonical not in imported_code_seen:
                imported_codes.append(canonical)
                imported_code_seen.add(canonical)

        for product_code, row in product_seed_rows.items():
            savepoint = f"inventory_product_{len(imported_codes) + 1}"
            connection.execute(f"SAVEPOINT {savepoint}")
            try:
                try:
                    validate_product_code(product_code)
                    validate_product_terms(product_code, row)
                except ValueError:
                    connection.execute(f"RELEASE SAVEPOINT {savepoint}")
                    continue
                now = utc_now_iso()
                use_canonical_product = should_use_canonical_product(product_code, row)
                product = find_product_by_code(connection, product_code) if use_canonical_product else find_existing_product(connection, product_code)
                if not product:
                    if use_canonical_product:
                        product = ensure_product(connection, product_code, row, now)
                        report["products_created"] += 1
                    else:
                        connection.execute(f"RELEASE SAVEPOINT {savepoint}")
                        continue
                elif use_canonical_product:
                    sync_product_metadata(connection, product["id"], product_code, row, now)
                    report["products_updated"] += 1
                add_imported_code(product_code)
                if mode == "replace":
                    for product_id in active_product_ids_for_code(connection, product_code):
                        if product_id not in replaced_product_ids:
                            report["credentials_disabled"] += disable_available_inventory(connection, product_id, now)
                            replaced_product_ids.add(product_id)
                connection.execute(f"RELEASE SAVEPOINT {savepoint}")
            except Exception as exc:
                connection.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                connection.execute(f"RELEASE SAVEPOINT {savepoint}")
                report["row_errors"] += 1
                report["invalid_rows"] += 1
                report["errors"].append(f"product {product_code}: {exc}")

        for row_index, (row_number, row) in enumerate(rows, start=1):
            savepoint = f"inventory_row_{row_index}"
            connection.execute(f"SAVEPOINT {savepoint}")
            try:
                product_code = normalize_import_product_code(text(row["product_code"]), row)
                row["product_code"] = product_code
                if not product_code:
                    raise ValueError("product_code is required")
                validate_product_code(product_code)
                validate_product_terms(product_code, row)
                credential = credential_from_row(row)
                now = utc_now_iso()
                use_canonical_product = should_use_canonical_product(product_code, row)
                product = find_product_by_code(connection, product_code) if use_canonical_product else find_existing_product(connection, product_code)
                if not product:
                    if use_canonical_product:
                        product = ensure_product(connection, product_code, row, now)
                        report["products_created"] += 1
                    else:
                        raise ValueError(f"Unknown product_code: {product_code}. Product must exist in products; import did not create it.")
                product_id = product["id"]

                item_id = str(uuid.uuid4())
                connection.execute(
                    "INSERT INTO inventory_items (id, product_id, secret_value, status, created_at) VALUES (?, ?, ?, 'available', ?)",
                    (item_id, product_id, credential, now),
                )
                connection.execute(
                    """
                    INSERT INTO inventory_movements
                        (id, inventory_item_id, action, source, created_at)
                    VALUES (?, ?, 'import', 'import_inventory', ?)
                    """,
                    (str(uuid.uuid4()), item_id, now),
                )
                connection.execute(f"RELEASE SAVEPOINT {savepoint}")
                report["credentials_added"] += 1
                report["valid_rows"] += 1
                add_imported_code(product_code)
            except Exception as exc:
                connection.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
                connection.execute(f"RELEASE SAVEPOINT {savepoint}")
                report["row_errors"] += 1
                report["invalid_rows"] += 1
                report["errors"].append(f"row {row_number}: {exc}")
        report["stock"] = {
            code: available_stock_for_code(connection, code)
            for code in imported_codes
        }
    return report


def print_report(report: dict[str, Any]) -> None:
    print(f"Products created: {report['products_created']}")
    print(f"Products updated: {report['products_updated']}")
    print(f"Rows read: {report['rows_read']}")
    print(f"Valid rows: {report['valid_rows']}")
    print(f"Credentials added: {report['credentials_added']}")
    print(f"Credentials duplicate: {report['credentials_duplicate']}")
    print(f"Rows with errors: {report['row_errors']}")
    print(f"Invalid rows: {report['invalid_rows']}")
    print("Stock by product_code:")
    for code, stock in report["stock"].items():
        print(f"  {code}: {stock}")
    for error in report["errors"]:
        print(f"Error: {error}", file=sys.stderr)


def self_test() -> None:
    with tempfile.TemporaryDirectory() as directory:
        root = Path(directory)
        database = root / "store.db"
        shutil.copy2(DEFAULT_DATABASE, database)
        source = root / "inventory.csv"
        rows = [
            REQUIRED_COLUMNS,
            ["GEMINI", "AI", "Gemini", "personal", "30D", "70000", "7", "gemini@example.com|pass", "test", "true"],
            ["CHATGPT", "AI", "ChatGPT", "personal", "30D", "199000", "7", "chatgpt@example.com|pass|2fa", "test", "true"],
            ["GROK", "AI", "Grok", "personal", "30D", "299000", "7", "grok@example.com|pass|2fa|recovery@example.com", "test", "true"],
        ]
        with source.open("w", encoding="utf-8", newline="") as handle:
            csv.writer(handle).writerows(rows)
        report = import_inventory(source, database)
        assert report["products_created"] == 0, report
        assert report["credentials_added"] == 3, report
        assert report["credentials_duplicate"] == 0, report
        assert report["row_errors"] == 0, report
        assert report["stock"] == {
            "GEMINI": 1,
            "CHATGPT": 1,
            "GROK": 1,
        }, report
        print_report(report)
        print("SELF-TEST: PASS")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import account inventory from .xlsx or .csv")
    parser.add_argument("input", nargs="?", type=Path, help="Excel or CSV source file")
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE, help="Path to store.db")
    parser.add_argument("--mode", choices=("append", "replace"), default="replace", help="append to stock or replace available stock for imported products")
    parser.add_argument("--self-test", action="store_true", help="Run isolated Gemini/ChatGPT/Grok import test")
    args = parser.parse_args()
    if args.self_test:
        self_test()
        return 0
    if not args.input:
        parser.error("input is required unless --self-test is used")
    report = import_inventory(args.input, args.database, args.mode)
    print_report(report)
    return 1 if report["row_errors"] else 0


if __name__ == "__main__":
    raise SystemExit(main())
