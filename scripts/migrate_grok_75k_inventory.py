from __future__ import annotations

import argparse
import hashlib
import shutil
import sqlite3
import uuid
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


OLD_PRODUCT_CODE = "GROK"
NEW_PRODUCT_CODE = "GROK_75K"
NEW_PRODUCT_NAME = "SUPERGROK AI"
NEW_PRICE_VND = 75000
NEW_DURATION = "30D"
NEW_WARRANTY_DAYS = 7
EXPECTED_CREDENTIAL_COUNT = 10
DEFAULT_DATABASE = Path("/var/data/store.db")
EMBEDDED_CREDENTIAL_FINGERPRINTS = (
    "039e658b4a8b153cf294d9d6d3cfba27b8de1ec400c3d6b49731d1edb3a409f5",
    "0d08d35c9bc124b0463e7340d8fc9facb4c60aae886be02fae4742717897db8e",
    "18e0ce3b2c3f4fdcf243f91fe14e2307636b235865f8293d867f0267a0a66184",
    "2471dd635aa6fd23dd1a79e5e8b0acf76543acc9b30bea75a17715f2de8becf0",
    "321206535feced5e105b3cb1bd59cc4538dff79987fdf0edc9f63f1cd335ce7b",
    "5b047d0ca5664f5323139d4478991d996bc921867af62398eccac2d8d9b5d8c7",
    "78e5a6866cab044ac36a0cc752dd111b9609071e713908d293a3d86c8b105959",
    "8b3bdc44acf6f2fa7ebe5517af4195fabdc1e25ae0107d80f4619bbfe58388a6",
    "8c8b00f6acffa56f7c89713728115b7c66310dafa75d04531f935974b60eb9fd",
    "c321c4e1aabb566595d923aa77b76adab4e90d2449b6f10922756b98727ee9c1",
)


@dataclass(frozen=True)
class CredentialRecord:
    row_number: int
    credential: str | None
    fingerprint: str
    masked: str


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Safely migrate the exact imported 75k SuperGrok credentials from "
            "GROK to GROK_75K. Dry-run by default."
        )
    )
    parser.add_argument(
        "--excel",
        type=Path,
        help="Optional path to import_GROK_ONLY_READY1_price75000.xlsx. Used only to verify embedded fingerprints.",
    )
    parser.add_argument("--database", type=Path, default=DEFAULT_DATABASE, help="Path to store.db")
    parser.add_argument("--yes", action="store_true", help="Apply the migration. Omit for dry-run.")
    return parser.parse_args()


def parse_int(value: Any) -> int:
    if value is None or str(value).strip() == "":
        return 0
    return int(float(str(value).strip()))


def normalized_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def fingerprint(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()


def mask_credential(value: str) -> str:
    value = value.strip()
    if len(value) <= 12:
        return "***"
    return f"{value[:3]}***{value[-8:]}"


def read_excel_credentials(path: Path) -> list[CredentialRecord]:
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        records: list[CredentialRecord] = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [normalized_text(cell) for cell in rows[0]]
            for row_number, row in enumerate(rows[1:], start=2):
                if not any(normalized_text(cell) for cell in row):
                    continue
                data = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
                product_code = normalized_text(data.get("product_code")).upper()
                price_vnd = parse_int(data.get("price_vnd"))
                duration = normalized_text(data.get("duration")).upper().replace(" ", "")
                warranty_days = parse_int(data.get("warranty_days"))
                credential = normalized_text(data.get("credential_text"))
                if product_code != OLD_PRODUCT_CODE:
                    raise ValueError(f"row {row_number}: expected product_code={OLD_PRODUCT_CODE}, got {product_code!r}")
                if price_vnd != NEW_PRICE_VND or duration != NEW_DURATION or warranty_days != NEW_WARRANTY_DAYS:
                    raise ValueError(
                        f"row {row_number}: expected price={NEW_PRICE_VND}, duration={NEW_DURATION}, "
                        f"warranty={NEW_WARRANTY_DAYS}; got price={price_vnd}, duration={duration}, "
                        f"warranty={warranty_days}"
                    )
                if not credential:
                    raise ValueError(f"row {row_number}: credential_text is empty")
                records.append(
                    CredentialRecord(
                        row_number=row_number,
                        credential=credential,
                        fingerprint=fingerprint(credential),
                        masked=mask_credential(credential),
                    )
                )
    finally:
        workbook.close()
    if len(records) != EXPECTED_CREDENTIAL_COUNT:
        raise ValueError(f"expected {EXPECTED_CREDENTIAL_COUNT} credential rows, found {len(records)}")
    duplicate_counts = Counter(record.credential for record in records)
    duplicates = [mask_credential(value) for value, count in duplicate_counts.items() if count > 1]
    if duplicates:
        raise ValueError(f"Excel contains duplicate credentials: {', '.join(duplicates)}")
    return records


def embedded_credentials() -> list[CredentialRecord]:
    if len(EMBEDDED_CREDENTIAL_FINGERPRINTS) != EXPECTED_CREDENTIAL_COUNT:
        raise ValueError(
            f"expected {EXPECTED_CREDENTIAL_COUNT} embedded fingerprints, "
            f"found {len(EMBEDDED_CREDENTIAL_FINGERPRINTS)}"
        )
    duplicate_counts = Counter(EMBEDDED_CREDENTIAL_FINGERPRINTS)
    duplicates = [value for value, count in duplicate_counts.items() if count > 1]
    if duplicates:
        raise ValueError(f"embedded fingerprints contain duplicates: {', '.join(duplicates)}")
    return [
        CredentialRecord(
            row_number=index,
            credential=None,
            fingerprint=value,
            masked="fingerprint-only",
        )
        for index, value in enumerate(EMBEDDED_CREDENTIAL_FINGERPRINTS, start=1)
    ]


def load_credential_records(excel_path: Path | None) -> list[CredentialRecord]:
    embedded = embedded_credentials()
    if excel_path is None:
        return embedded
    excel_records = read_excel_credentials(excel_path)
    embedded_fingerprints = {record.fingerprint for record in embedded}
    excel_fingerprints = {record.fingerprint for record in excel_records}
    if excel_fingerprints != embedded_fingerprints:
        missing = sorted(embedded_fingerprints - excel_fingerprints)
        extra = sorted(excel_fingerprints - embedded_fingerprints)
        raise ValueError(
            "Excel fingerprints do not match embedded fingerprints: "
            f"missing={[value[:12] for value in missing]} extra={[value[:12] for value in extra]}"
        )
    return excel_records


def connect(database: Path) -> sqlite3.Connection:
    if not database.is_file():
        raise FileNotFoundError(f"store.db not found: {database}")
    connection = sqlite3.connect(database)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def table_columns(connection: sqlite3.Connection, table: str) -> set[str]:
    return {str(row["name"]) for row in connection.execute(f"PRAGMA table_info({table})")}


def find_product(connection: sqlite3.Connection, code: str) -> sqlite3.Row | None:
    return connection.execute("SELECT * FROM products WHERE UPPER(code) = ?", (code.upper(),)).fetchone()


def ensure_grok_75k_product(connection: sqlite3.Connection, now: str) -> str:
    columns = table_columns(connection, "products")
    existing = find_product(connection, NEW_PRODUCT_CODE)
    if existing:
        values: dict[str, Any] = {
            "name": NEW_PRODUCT_NAME,
            "active": 1,
            "delivery_type": "account",
            "updated_at": now,
            "category": "account",
            "account_type": "personal",
            "duration": NEW_DURATION,
            "price_vnd": NEW_PRICE_VND,
            "warranty_days": NEW_WARRANTY_DAYS,
            "category_key": NEW_PRODUCT_CODE,
            "product_group": "account",
            "show_in_menu": 1,
            "description": NEW_PRODUCT_NAME,
        }
        update_columns = [name for name in values if name in columns]
        assignments = ", ".join(f"{name} = ?" for name in update_columns)
        connection.execute(
            f"UPDATE products SET {assignments} WHERE id = ?",
            [values[name] for name in update_columns] + [existing["id"]],
        )
        return str(existing["id"])

    product_id = str(uuid.uuid4())
    values: dict[str, Any] = {
        "id": product_id,
        "code": NEW_PRODUCT_CODE,
        "name": NEW_PRODUCT_NAME,
        "active": 1,
        "delivery_type": "account",
        "created_at": now,
        "updated_at": now,
        "category": "account",
        "account_type": "personal",
        "duration": NEW_DURATION,
        "price_vnd": NEW_PRICE_VND,
        "warranty_days": NEW_WARRANTY_DAYS,
        "note": "",
        "menu_order": 100,
        "show_in_menu": 1,
        "product_group": "account",
        "category_key": NEW_PRODUCT_CODE,
        "description": NEW_PRODUCT_NAME,
    }
    insert_columns = [name for name in values if name in columns]
    placeholders = ", ".join("?" for _ in insert_columns)
    connection.execute(
        f"INSERT INTO products ({', '.join(insert_columns)}) VALUES ({placeholders})",
        [values[name] for name in insert_columns],
    )
    return product_id


def fetch_matches(connection: sqlite3.Connection, records: list[CredentialRecord]) -> dict[str, list[sqlite3.Row]]:
    matches: dict[str, list[sqlite3.Row]] = defaultdict(list)
    wanted = {record.fingerprint for record in records}
    for row in connection.execute(
        """
        SELECT
            i.id AS inventory_item_id,
            i.product_id,
            p.code AS product_code,
            i.secret_value,
            i.status,
            i.reserved_order_id,
            i.delivered_order_id,
            i.created_at,
            i.reserved_at,
            i.delivered_at,
            i.disabled_at
        FROM inventory_items AS i
        JOIN products AS p ON p.id = i.product_id
        ORDER BY p.code, i.status, i.created_at, i.id
        """
    ):
        row_fingerprint = fingerprint(str(row["secret_value"]))
        if row_fingerprint in wanted:
            matches[row_fingerprint].append(row)
    return matches


def eligible_rows(
    records: list[CredentialRecord],
    matches: dict[str, list[sqlite3.Row]],
) -> list[sqlite3.Row]:
    rows: list[sqlite3.Row] = []
    for record in records:
        for row in matches[record.fingerprint]:
            if (
                row["product_code"] == OLD_PRODUCT_CODE
                and row["status"] == "available"
                and row["reserved_order_id"] is None
                and row["delivered_order_id"] is None
            ):
                rows.append(row)
    return rows


def available_count(connection: sqlite3.Connection, product_code: str) -> int:
    row = connection.execute(
        """
        SELECT COUNT(*) AS count
        FROM inventory_items AS i
        JOIN products AS p ON p.id = i.product_id
        WHERE p.code = ? AND i.status = 'available'
        """,
        (product_code,),
    ).fetchone()
    return int(row["count"] or 0)


def print_report(
    connection: sqlite3.Connection,
    records: list[CredentialRecord],
    matches: dict[str, list[sqlite3.Row]],
    eligible: list[sqlite3.Row],
) -> None:
    print(f"Embedded fingerprints count: {len(EMBEDDED_CREDENTIAL_FINGERPRINTS)}")
    print(f"Credential fingerprints in use: {len(records)}")
    print(f"Distinct credential fingerprints: {len({record.fingerprint for record in records})}")
    for record in records:
        rows = matches[record.fingerprint]
        print(f"- source_row={record.row_number} sha256={record.fingerprint[:12]} credential={record.masked} matched_rows={len(rows)}")
        for row in rows:
            print(
                "  "
                f"inventory_item_id={row['inventory_item_id']} "
                f"product_code={row['product_code']} "
                f"status={row['status']} "
                f"reserved_order_id={row['reserved_order_id']} "
                f"delivered_order_id={row['delivered_order_id']}"
            )
    print(f"Eligible migration rows ({OLD_PRODUCT_CODE} available only): {len(eligible)}")
    print(f"Before {OLD_PRODUCT_CODE} available: {available_count(connection, OLD_PRODUCT_CODE)}")
    print(f"Before {NEW_PRODUCT_CODE} available: {available_count(connection, NEW_PRODUCT_CODE)}")


def backup_database(database: Path) -> Path:
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_path = database.with_name(f"{database.name}.bak_before_grok_75k_migration_{timestamp}")
    shutil.copy2(database, backup_path)
    return backup_path


def apply_migration(database: Path, records: list[CredentialRecord]) -> None:
    backup_path = backup_database(database)
    print(f"Backup created: {backup_path}")
    with connect(database) as connection:
        connection.execute("BEGIN IMMEDIATE")
        try:
            now = utc_now_iso()
            product_id = ensure_grok_75k_product(connection, now)
            matches = fetch_matches(connection, records)
            eligible = eligible_rows(records, matches)
            if len(eligible) != EXPECTED_CREDENTIAL_COUNT:
                raise RuntimeError(
                    f"ABORT: expected exactly {EXPECTED_CREDENTIAL_COUNT} eligible rows, found {len(eligible)}"
                )
            for row in eligible:
                cursor = connection.execute(
                    """
                    UPDATE inventory_items
                    SET product_id = ?
                    WHERE id = ?
                      AND product_id = ?
                      AND status = 'available'
                      AND reserved_order_id IS NULL
                      AND delivered_order_id IS NULL
                    """,
                    (product_id, row["inventory_item_id"], row["product_id"]),
                )
                if cursor.rowcount != 1:
                    raise RuntimeError(f"ABORT: failed to update inventory item {row['inventory_item_id']}")
            if available_count(connection, NEW_PRODUCT_CODE) != EXPECTED_CREDENTIAL_COUNT:
                raise RuntimeError(
                    f"ABORT: expected {NEW_PRODUCT_CODE} available={EXPECTED_CREDENTIAL_COUNT}, "
                    f"got {available_count(connection, NEW_PRODUCT_CODE)}"
                )
            connection.commit()
        except Exception:
            connection.rollback()
            raise


def main() -> int:
    args = parse_args()
    records = load_credential_records(args.excel)
    with connect(args.database) as connection:
        matches = fetch_matches(connection, records)
        eligible = eligible_rows(records, matches)
        print_report(connection, records, matches, eligible)
        if not args.yes:
            print("DRY RUN ONLY: no database changes written. Re-run with --yes to apply.")
            return 0
        if len(eligible) != EXPECTED_CREDENTIAL_COUNT:
            raise RuntimeError(
                f"ABORT: expected exactly {EXPECTED_CREDENTIAL_COUNT} eligible rows, found {len(eligible)}"
            )
    apply_migration(args.database, records)
    with connect(args.database) as connection:
        print(f"After {OLD_PRODUCT_CODE} available: {available_count(connection, OLD_PRODUCT_CODE)}")
        print(f"After {NEW_PRODUCT_CODE} available: {available_count(connection, NEW_PRODUCT_CODE)}")
        product = find_product(connection, NEW_PRODUCT_CODE)
        if product:
            print(
                f"{NEW_PRODUCT_CODE}: name={product['name']} price_vnd={product['price_vnd']} "
                f"duration={product['duration']} warranty_days={product['warranty_days']} active={product['active']}"
            )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
