"""Disable expired GROK inventory from an Excel file on Render.

Safe default for Render Shell:
    python scripts/disable_expired_grok_render.py

Optional:
    python scripts/disable_expired_grok_render.py --database /var/data/store.db --excel imports/import_GROK_ONLY_READY1_price10000.xlsx

This script never prints full credentials and never deletes inventory rows.
It only changes matching GROK inventory_items that are currently status='available'.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import sqlite3
import sys
import uuid
from contextlib import closing
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.import_inventory import (  # noqa: E402
    SKIPPED_SHEET_NAMES,
    canonical_product_code,
    credential_from_row,
    normalize_import_product_code,
)


DEFAULT_DATABASE = Path("/var/data/store.db")
DEFAULT_EXCEL_NAME = "import_GROK_ONLY_READY1_price10000.xlsx"


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def fingerprint(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]


def masked_hint(value: str) -> str:
    first = value.split("|", 1)[0]
    if "@" in first:
        local, domain = first.split("@", 1)
        return f"{local[:2]}***@{domain[:2]}***"
    return f"{first[:2]}***"


def resolve_excel_path(raw: str | None) -> Path:
    candidates: list[Path] = []
    if raw:
        candidates.append(Path(raw).expanduser())
    env_path = os.environ.get("EXPIRED_GROK_XLSX", "").strip()
    if env_path:
        candidates.append(Path(env_path).expanduser())
    candidates.extend(
        [
            PROJECT_ROOT / "imports" / DEFAULT_EXCEL_NAME,
            Path.cwd() / "imports" / DEFAULT_EXCEL_NAME,
            Path("/var/data") / "imports" / DEFAULT_EXCEL_NAME,
            Path("/var/data") / DEFAULT_EXCEL_NAME,
        ]
    )
    for candidate in candidates:
        path = candidate if candidate.is_absolute() else PROJECT_ROOT / candidate
        if path.is_file():
            return path.resolve()
    searched = "\n".join(str((p if p.is_absolute() else PROJECT_ROOT / p).resolve()) for p in candidates)
    raise FileNotFoundError(f"Excel file not found. Searched:\n{searched}")


def read_grok_credentials(excel_path: Path) -> list[str]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    credentials: list[str] = []
    for sheet in workbook.worksheets:
        if sheet.title.strip().upper() in SKIPPED_SHEET_NAMES:
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        for row_index, values in enumerate(rows[1:], start=2):
            if not values or all(value is None or str(value).strip() == "" for value in values):
                continue
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers)) if headers[i]}
            product_code = canonical_product_code(normalize_import_product_code(str(row.get("product_code") or ""), row))
            if product_code != "GROK":
                raise RuntimeError(f"Unexpected product_code={product_code!r} at sheet={sheet.title!r} row={row_index}")
            credentials.append(credential_from_row(row))
    unique = list(dict.fromkeys(credentials))
    if len(unique) != len(credentials):
        raise RuntimeError(f"Excel contains duplicate credential rows: rows={len(credentials)} unique={len(unique)}")
    if not unique:
        raise RuntimeError("No GROK credentials found in Excel file")
    return unique


def backup_database(database_path: Path) -> Path:
    if not database_path.is_file():
        raise FileNotFoundError(f"Database not found: {database_path}")
    backup_dir = database_path.parent / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    backup_path = backup_dir / f"store_before_disable_expired_grok_{timestamp}.db"
    shutil.copy2(database_path, backup_path)
    return backup_path


def grok_stock_counts(connection: sqlite3.Connection) -> list[dict[str, Any]]:
    rows = connection.execute(
        """
        SELECT p.code, i.status, COUNT(*) AS count
        FROM products AS p
        JOIN inventory_items AS i ON i.product_id = p.id
        WHERE UPPER(p.code) LIKE '%GROK%' OR UPPER(p.name) LIKE '%GROK%'
        GROUP BY p.code, i.status
        ORDER BY p.code, i.status
        """
    ).fetchall()
    return [dict(row) for row in rows]


def available_grok_count(connection: sqlite3.Connection) -> int:
    row = connection.execute(
        """
        SELECT COUNT(*) AS count
        FROM inventory_items AS i
        JOIN products AS p ON p.id = i.product_id
        WHERE p.code = 'GROK' AND i.status = 'available'
        """
    ).fetchone()
    return int(row["count"] or 0)


def placeholders(values: list[str]) -> str:
    return ",".join("?" for _ in values)


def disable_matching_available(database_path: Path, credentials: list[str]) -> dict[str, Any]:
    now = utc_now_iso()
    with closing(sqlite3.connect(database_path)) as connection:
        connection.row_factory = sqlite3.Row
        before_counts = grok_stock_counts(connection)
        before_available = available_grok_count(connection)
        try:
            connection.execute("BEGIN IMMEDIATE")
            matched_rows = connection.execute(
                f"""
                SELECT i.id, i.secret_value, i.status, i.reserved_order_id, i.delivered_order_id,
                       i.reserved_at, i.delivered_at, p.code AS product_code, p.name AS product_name
                FROM inventory_items AS i
                JOIN products AS p ON p.id = i.product_id
                WHERE p.code = 'GROK'
                  AND i.secret_value IN ({placeholders(credentials)})
                ORDER BY i.created_at, i.id
                """,
                credentials,
            ).fetchall()
            available_rows = [row for row in matched_rows if row["status"] == "available"]
            blocked_rows = []
            for row in available_rows:
                linked_orders = connection.execute(
                    """
                    SELECT COUNT(*) AS count
                    FROM order_inventory_items
                    WHERE inventory_item_id = ?
                    """,
                    (row["id"],),
                ).fetchone()["count"]
                if row["reserved_order_id"] or row["delivered_order_id"] or row["reserved_at"] or row["delivered_at"] or int(linked_orders or 0):
                    blocked_rows.append(row)
            if blocked_rows:
                blocked = ", ".join(fingerprint(row["secret_value"]) for row in blocked_rows)
                raise RuntimeError(f"Rollback: available matched rows have reservation/order evidence: {blocked}")

            item_ids = [row["id"] for row in available_rows]
            if item_ids:
                connection.executemany(
                    """
                    UPDATE inventory_items
                    SET status = 'disabled', disabled_at = ?
                    WHERE id = ? AND status = 'available'
                    """,
                    [(now, item_id) for item_id in item_ids],
                )
                connection.executemany(
                    """
                    INSERT INTO inventory_movements
                        (id, inventory_item_id, action, order_id, admin_telegram_id, source, created_at)
                    VALUES (?, ?, 'disable', NULL, NULL, 'render_disable_expired_grok_excel', ?)
                    """,
                    [(str(uuid.uuid4()), item_id, now) for item_id in item_ids],
                )

            remaining_available_matches = connection.execute(
                f"""
                SELECT COUNT(*) AS count
                FROM inventory_items AS i
                JOIN products AS p ON p.id = i.product_id
                WHERE p.code = 'GROK'
                  AND i.status = 'available'
                  AND i.secret_value IN ({placeholders(credentials)})
                """,
                credentials,
            ).fetchone()["count"]
            if int(remaining_available_matches or 0) != 0:
                raise RuntimeError(f"Rollback: {remaining_available_matches} matched expired credentials would remain available")

            after_counts = grok_stock_counts(connection)
            after_available = available_grok_count(connection)
            connection.commit()
            return {
                "before_counts": before_counts,
                "after_counts": after_counts,
                "before_available": before_available,
                "after_available": after_available,
                "matched_total": len(matched_rows),
                "matched_available": len(available_rows),
                "disabled_count": len(item_ids),
                "matched_status_counts": status_counts(matched_rows),
                "disabled_fingerprints": [fingerprint(row["secret_value"]) for row in available_rows],
                "disabled_hints": [masked_hint(row["secret_value"]) for row in available_rows],
                "disabled_at_utc": now,
            }
        except Exception:
            connection.rollback()
            raise


def status_counts(rows: list[sqlite3.Row]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        status = str(row["status"])
        counts[status] = counts.get(status, 0) + 1
    return counts


def live_menu_count(database_path: Path) -> int:
    os.environ["STORE_DB_PATH"] = str(database_path)
    import telegram_license_bot

    return int(telegram_license_bot._menu_available_count("GROK"))


def main() -> int:
    parser = argparse.ArgumentParser(description="Disable expired GROK credentials on Render safely.")
    parser.add_argument("--database", default=str(DEFAULT_DATABASE), help="Production DB path. Default: /var/data/store.db")
    parser.add_argument("--excel", default=None, help="Expired GROK Excel path. Default: auto-detect imports/import_GROK_ONLY_READY1_price10000.xlsx")
    parser.add_argument("--yes", action="store_true", help="Required confirmation flag.")
    args = parser.parse_args()

    database_path = Path(args.database).expanduser()
    if not database_path.is_absolute():
        database_path = (PROJECT_ROOT / database_path).resolve()
    excel_path = resolve_excel_path(args.excel)

    print("SAFE GROK DISABLE SCRIPT")
    print(f"DATABASE: {database_path.resolve()}")
    print(f"EXCEL: {excel_path}")
    print("MODE: disable only matching GROK inventory_items where status='available'")
    print("NO FULL CREDENTIALS WILL BE PRINTED")
    if database_path.resolve() != DEFAULT_DATABASE:
        print("WARNING: database is not /var/data/store.db")
    if not args.yes:
        print("DRY-RUN ONLY: rerun with --yes to create backup and commit changes.")
        credentials = read_grok_credentials(excel_path)
        with closing(sqlite3.connect(database_path)) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                f"""
                SELECT i.secret_value, i.status
                FROM inventory_items AS i
                JOIN products AS p ON p.id = i.product_id
                WHERE p.code = 'GROK'
                  AND i.secret_value IN ({placeholders(credentials)})
                """,
                credentials,
            ).fetchall()
            print(f"EXCEL_GROK_CREDENTIALS: {len(credentials)}")
            print(f"MATCHED_DB_ROWS: {len(rows)}")
            print(f"MATCHED_STATUS_COUNTS: {status_counts(rows)}")
            print(f"GROK_STOCK_BEFORE: {grok_stock_counts(connection)}")
        return 2

    credentials = read_grok_credentials(excel_path)
    backup_path = backup_database(database_path)
    print(f"BACKUP_CREATED: {backup_path}")
    result = disable_matching_available(database_path, credentials)
    print(f"EXCEL_GROK_CREDENTIALS: {len(credentials)}")
    print(f"MATCHED_DB_ROWS: {result['matched_total']}")
    print(f"MATCHED_STATUS_COUNTS: {result['matched_status_counts']}")
    print(f"MATCHED_AVAILABLE_TO_DISABLE: {result['matched_available']}")
    print(f"DISABLED_COUNT: {result['disabled_count']}")
    print(f"DISABLED_FINGERPRINTS: {', '.join(result['disabled_fingerprints'])}")
    print(f"DISABLED_SAFE_HINTS: {', '.join(result['disabled_hints'])}")
    print(f"DISABLED_AT_UTC: {result['disabled_at_utc']}")
    print(f"GROK_STOCK_BEFORE: {result['before_counts']}")
    print(f"GROK_AVAILABLE_BEFORE: {result['before_available']}")
    print(f"GROK_STOCK_AFTER: {result['after_counts']}")
    print(f"GROK_AVAILABLE_AFTER: {result['after_available']}")
    print(f'LIVE_MENU_HELPER__menu_available_count("GROK"): {live_menu_count(database_path)}')
    print("DONE")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
