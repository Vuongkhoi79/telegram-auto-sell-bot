from __future__ import annotations

import sqlite3
import sys
import tempfile
import unittest
import csv
import os
from contextlib import closing
from datetime import datetime, timedelta, timezone
from pathlib import Path

from openpyxl import Workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from repository.store_repository import StoreRepository
from scripts.import_inventory import OPTIONAL_COLUMNS, REQUIRED_COLUMNS, import_inventory
from scripts.cleanup_demo_inventory import cleanup_demo_inventory
import telegram_license_bot as bot


class StoreRepositoryTest(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = tempfile.TemporaryDirectory()
        self.db_path = Path(self.temp_dir.name) / "store.db"
        source_db = PROJECT_ROOT / "database" / "store.db"
        with closing(sqlite3.connect(source_db)) as source, closing(sqlite3.connect(self.db_path)) as target:
            source.backup(target)
        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                connection.execute(
                    "INSERT INTO products (id, code, name, active, delivery_type, created_at, updated_at) VALUES (?, ?, ?, 1, 'account', ?, ?)",
                    ("product-test", "TEST_PRODUCT", "Test Product", now, now),
                )
                connection.execute(
                    """
                    INSERT INTO orders
                        (id, order_id, telegram_user_id, product_code, product_name, package_name, quantity,
                         unit_price_vnd, total_vnd, delivery_type, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("order-test", "ORD-TEST-001", 1, "TEST_PRODUCT", "Test Product", "TEST", 1, 1, 1, "account", now),
                )
        self.store = StoreRepository(self.db_path)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_reserve_deliver_and_duplicate_transaction(self) -> None:
        self.assertIn(
            "TEST_PRODUCT", {product["code"] for product in self.store.list_active_products()}
        )
        self.store.add_inventory_item("TEST_PRODUCT", "first@example.com|pass1")
        self.store.add_inventory_item("TEST_PRODUCT", "second@example.com|pass2")
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 2)

        reserved = self.store.reserve_inventory_items("ORD-TEST-001", "TEST_PRODUCT", 1)
        self.assertEqual(len(reserved), 1)
        reserved_credential = reserved[0]["credential_text"]
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 1)

        self.assertTrue(self.store.mark_order_paid("ORD-TEST-001", "SEPAY-TEST-001"))
        delivered = self.store.deliver_reserved_items("ORD-TEST-001")
        self.assertEqual(delivered, [reserved_credential])
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 1)

        with closing(sqlite3.connect(self.db_path)) as connection:
            status = connection.execute(
                "SELECT status FROM inventory_items WHERE secret_value = ?", (reserved_credential,)
            ).fetchone()[0]
        self.assertEqual(status, "delivered")
        self.assertFalse(self.store.mark_order_paid("ORD-TEST-001", "SEPAY-TEST-001"))
        self.assertEqual(self.store.deliver_reserved_items("ORD-TEST-001"), [reserved_credential])

        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                connection.execute(
                    """
                    INSERT INTO orders
                        (id, order_id, telegram_user_id, product_code, product_name, package_name, quantity,
                         unit_price_vnd, total_vnd, delivery_type, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    ("order-test-2", "ORD-TEST-002", 1, "TEST_PRODUCT", "Test Product", "TEST", 2, 1, 2, "account", now),
                )
        with self.assertRaisesRegex(ValueError, "Insufficient available inventory"):
            self.store.reserve_inventory_items("ORD-TEST-002", "TEST_PRODUCT", 2)

    def test_canonical_account_product_codes_resolve_legacy_product_rows(self) -> None:
        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        initial_stock = self.store.get_stock_count("CHATGPT")
        item_id = self.store.add_inventory_item("CHATGPT", "legacy@example.com|pass")
        self.assertTrue(item_id)
        self.assertEqual(self.store.get_stock_count("CHATGPT"), initial_stock + 1)
        reserved = self.store.create_pending_account_order_and_reserve(
            order_id="ORD-LEGACY-CHATGPT",
            telegram_user_id=1,
            username="Legacy Test",
            product_code="CHATGPT",
            product_name="CHATGPT",
            package_name="7D",
            quantity=1,
            unit_price_vnd=1,
            total_vnd=1,
            created_at=now,
            expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        )
        self.assertEqual(len(reserved), 1)
        with closing(sqlite3.connect(self.db_path)) as connection:
            row = connection.execute(
                "SELECT product_code, product_name FROM orders WHERE order_id = ?",
                ("ORD-LEGACY-CHATGPT",),
            ).fetchone()
        self.assertEqual((row[0], row[1]), ("CHATGPT", "CHATGPT"))

    def test_get_orders_by_telegram_user_returns_latest_ten_only_for_that_user(self) -> None:
        now = datetime.now(timezone.utc).replace(microsecond=0)
        with closing(sqlite3.connect(self.db_path)) as connection, connection:
            for index in range(12):
                created_at = (now + timedelta(minutes=index)).isoformat()
                connection.execute(
                    """
                    INSERT INTO orders
                        (id, order_id, telegram_user_id, product_code, product_name, package_name, quantity,
                         unit_price_vnd, total_vnd, delivery_type, payment_status, order_status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        f"order-history-{index}",
                        f"ORD-HISTORY-{index:02d}",
                        42,
                        "GEMINI",
                        "Gemini AI Pro",
                        "Gemini AI Pro",
                        1,
                        70000,
                        70000,
                        "account",
                        "paid" if index % 2 else "pending",
                        "delivered" if index % 3 == 0 else "pending",
                        created_at,
                    ),
                )
            connection.execute(
                """
                INSERT INTO orders
                    (id, order_id, telegram_user_id, product_code, product_name, package_name, quantity,
                     unit_price_vnd, total_vnd, delivery_type, payment_status, order_status, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    "order-history-other",
                    "ORD-HISTORY-OTHER",
                    7,
                    "CAPCUT",
                    "CAPCUT PRO",
                    "CAPCUT PRO",
                    1,
                    400000,
                    400000,
                    "account",
                    "paid",
                    "delivered",
                    (now + timedelta(minutes=100)).isoformat(),
                ),
            )
        orders = self.store.get_orders_by_telegram_user(42, limit=10)
        self.assertEqual(len(orders), 10)
        self.assertEqual(orders[0]["order_id"], "ORD-HISTORY-11")
        self.assertEqual(orders[-1]["order_id"], "ORD-HISTORY-02")
        self.assertTrue(all(order["telegram_user_id"] == 42 for order in orders))
        self.assertTrue(all(order["order_id"] != "ORD-HISTORY-OTHER" for order in orders))

    def test_release_expired_reservations_preserves_paid_and_delivered_items(self) -> None:
        for credential in ("expired@example.com|pass", "paid@example.com|pass", "delivered@example.com|pass"):
            self.store.add_inventory_item("TEST_PRODUCT", credential)
        now = datetime.now(timezone.utc).replace(microsecond=0)
        expired_at = (now - timedelta(minutes=1)).isoformat()
        created_at = (now - timedelta(minutes=6)).isoformat()

        def reserve(order_id: str) -> None:
            self.store.create_pending_account_order_and_reserve(
                order_id=order_id,
                telegram_user_id=1,
                username="Test User",
                product_code="TEST_PRODUCT",
                product_name="Test Product",
                package_name="TEST",
                quantity=1,
                unit_price_vnd=1,
                total_vnd=1,
                created_at=created_at,
                expire_at=expired_at,
            )

        reserve("ORD-EXPIRED")
        reserve("ORD-PAID")
        reserve("ORD-DELIVERED")
        self.store.mark_account_order_paid_for_fulfillment("ORD-PAID")
        self.store.mark_account_order_paid_for_fulfillment("ORD-DELIVERED")
        self.assertEqual(len(self.store.deliver_reserved_items("ORD-DELIVERED")), 1)

        self.assertEqual(self.store.release_expired_reservations(5), 1)
        with closing(sqlite3.connect(self.db_path)) as connection:
            order_items = {
                row[0]: (row[1], row[2], row[3])
                for row in connection.execute(
                    """
                    SELECT orders.order_id, inventory_items.status,
                           COALESCE(inventory_items.reserved_order_id, ''), order_inventory_items.state
                    FROM order_inventory_items
                    JOIN orders ON orders.id = order_inventory_items.order_id
                    JOIN inventory_items ON inventory_items.id = order_inventory_items.inventory_item_id
                    """
                )
            }
        self.assertNotIn("ORD-EXPIRED", order_items)
        self.assertEqual(order_items["ORD-PAID"][0], "reserved")
        self.assertEqual(order_items["ORD-DELIVERED"][0], "delivered")
        self.assertEqual(self.store.release_expired_reservations(5), 0)

    def test_release_order_reservation_returns_stock_before_payment(self) -> None:
        for credential in ("release1@example.com|pass", "release2@example.com|pass"):
            self.store.add_inventory_item("TEST_PRODUCT", credential)
        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        self.store.create_pending_account_order_and_reserve(
            order_id="ORD-RELEASE-001",
            telegram_user_id=1,
            username="Test User",
            product_code="TEST_PRODUCT",
            product_name="Test Product",
            package_name="TEST",
            quantity=2,
            unit_price_vnd=1,
            total_vnd=2,
            created_at=now,
            expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        )
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 0)
        released = self.store.release_order_reservation("ORD-RELEASE-001")
        self.assertEqual(released, 2)
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 2)
        with closing(sqlite3.connect(self.db_path)) as connection:
            rows = dict(connection.execute(
                "SELECT secret_value, status FROM inventory_items WHERE secret_value LIKE 'release%@example.com|pass'"
            ).fetchall())
            link_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM order_inventory_items
                JOIN orders ON orders.id = order_inventory_items.order_id
                JOIN inventory_items ON inventory_items.id = order_inventory_items.inventory_item_id
                WHERE orders.order_id = 'ORD-RELEASE-001'
                """
            ).fetchone()[0]
        self.assertEqual(set(rows.values()), {"available"})
        self.assertEqual(link_count, 0)

    def test_paid_order_is_not_released_by_timeout_and_still_delivers(self) -> None:
        for credential in ("paid-timeout1@example.com|pass", "paid-timeout2@example.com|pass"):
            self.store.add_inventory_item("TEST_PRODUCT", credential)
        now = datetime.now(timezone.utc).replace(microsecond=0)
        order_id = "ORD-PAID-TIMEOUT"
        self.store.create_pending_account_order_and_reserve(
            order_id=order_id,
            telegram_user_id=1,
            username="Test User",
            product_code="TEST_PRODUCT",
            product_name="Test Product",
            package_name="TEST",
            quantity=2,
            unit_price_vnd=1,
            total_vnd=2,
            created_at=now.isoformat(),
            expire_at=(now + timedelta(minutes=5)).isoformat(),
        )
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                connection.execute(
                    "UPDATE inventory_items SET reserved_at = ? WHERE reserved_order_id = ?",
                    ((now - timedelta(minutes=6)).isoformat(), order_id),
                )
        self.store.mark_account_order_paid_for_fulfillment(order_id)
        self.assertEqual(self.store.release_expired_reservations(5), 0)
        delivered = self.store.deliver_reserved_items(order_id)
        self.assertEqual(len(delivered), 2)
        self.assertEqual(self.store.get_stock_count("TEST_PRODUCT"), 0)
        with closing(sqlite3.connect(self.db_path)) as connection:
            link_count = connection.execute(
                """
                SELECT COUNT(*)
                FROM order_inventory_items
                JOIN orders ON orders.id = order_inventory_items.order_id
                WHERE orders.order_id = ?
                """,
                (order_id,),
            ).fetchone()[0]
        self.assertEqual(link_count, 2)

    def test_stale_link_is_cleaned_before_reservation(self) -> None:
        self.store.add_inventory_item("TEST_PRODUCT", "stale@example.com|pass")
        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                stale_order_id = "ORD-STALE-001"
                connection.execute(
                    """
                    INSERT INTO orders
                        (id, order_id, telegram_user_id, product_code, product_name, package_name, quantity,
                         unit_price_vnd, total_vnd, delivery_type, created_at, payment_status, order_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pending', 'reserved')
                    """,
                    ("order-stale", stale_order_id, 1, "TEST_PRODUCT", "Test Product", "TEST", 1, 1, 1, "account", now),
                )
                stale_item_id = connection.execute(
                    "SELECT id FROM inventory_items WHERE secret_value = 'stale@example.com|pass'"
                ).fetchone()[0]
                connection.execute(
                    """
                    INSERT INTO order_inventory_items (order_id, inventory_item_id, state, created_at)
                    VALUES (?, ?, 'reserved', ?)
                    """,
                    ("order-stale", stale_item_id, now),
                )
        reserved = self.store.create_pending_account_order_and_reserve(
            order_id="ORD-STALE-CLEANUP",
            telegram_user_id=1,
            username="Stale Test",
            product_code="TEST_PRODUCT",
            product_name="Test Product",
            package_name="TEST",
            quantity=1,
            unit_price_vnd=1,
            total_vnd=1,
            created_at=now,
            expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        )
        self.assertEqual(len(reserved), 1)
        with closing(sqlite3.connect(self.db_path)) as connection:
            link_rows = connection.execute(
                "SELECT COUNT(*) FROM order_inventory_items WHERE inventory_item_id = ?",
                (stale_item_id,),
            ).fetchone()[0]
        self.assertEqual(link_rows, 1)

    def test_admin_stock_operations_and_repeat_import(self) -> None:
        item_ids = [
            self.store.add_inventory_item("TEST_PRODUCT", credential)
            for credential in ("one@example.com|pass", "two@example.com|pass", "three@example.com|pass")
        ]
        with self.assertRaisesRegex(ValueError, "Duplicate credential"):
            self.store.add_inventory_item("TEST_PRODUCT", "one@example.com|pass")

        self.store.set_inventory_item_disabled(item_ids[0], True)
        self.store.set_product_active("TEST_PRODUCT", False)
        with self.assertRaisesRegex(ValueError, "product is disabled"):
            self.store.set_inventory_item_disabled(item_ids[0], False)
        self.store.set_product_active("TEST_PRODUCT", True)
        self.store.set_inventory_item_disabled(item_ids[0], False)

        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        reserved_ids = self.store.create_pending_account_order_and_reserve(
            order_id="ORD-ADMIN-RESERVED",
            telegram_user_id=1,
            username="Admin Test",
            product_code="TEST_PRODUCT",
            product_name="Test Product",
            package_name="TEST",
            quantity=1,
            unit_price_vnd=1,
            total_vnd=1,
            created_at=now,
            expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        )
        delivered_ids = self.store.create_pending_account_order_and_reserve(
            order_id="ORD-ADMIN-DELIVERED",
            telegram_user_id=1,
            username="Admin Test",
            product_code="TEST_PRODUCT",
            product_name="Test Product",
            package_name="TEST",
            quantity=1,
            unit_price_vnd=1,
            total_vnd=1,
            created_at=now,
            expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        )
        self.store.mark_account_order_paid_for_fulfillment("ORD-ADMIN-DELIVERED")
        self.assertEqual(len(self.store.deliver_reserved_items("ORD-ADMIN-DELIVERED")), 1)
        remaining_id = (set(item_ids) - set(reserved_ids) - set(delivered_ids)).pop()
        self.store.set_inventory_item_disabled(remaining_id, True)
        with self.assertRaisesRegex(ValueError, "Only available"):
            self.store.set_inventory_item_disabled(delivered_ids[0], True)

        summary = {row["product_code"]: row for row in self.store.get_stock_summary()}["TEST_PRODUCT"]
        self.assertEqual(
            (summary["active"], summary["available"], summary["reserved"], summary["delivered"], summary["disabled"]),
            (1, 0, 1, 1, 1),
        )

        import_path = Path(self.temp_dir.name) / "inventory.csv"
        with import_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(REQUIRED_COLUMNS)
            writer.writerow(["GEMINI", "AI", "Gemini", "private", "30D", 100, 7, "import1@example.com|pass", "", 1])
            writer.writerow(["GEMINI", "AI", "Gemini", "private", "30D", 100, 7, "import2@example.com|pass|2fa", "", 1])
        first_import = import_inventory(import_path, self.db_path)
        second_import = import_inventory(import_path, self.db_path, mode="append")
        self.assertEqual((first_import["credentials_added"], first_import["credentials_duplicate"], first_import["row_errors"]), (2, 0, 0))
        self.assertEqual((second_import["credentials_added"], second_import["credentials_duplicate"], second_import["row_errors"]), (2, 0, 0))
        self.assertEqual(self.store.get_stock_count("GEMINI"), 4)

        row_isolation_path = Path(self.temp_dir.name) / "row-isolation.csv"
        with row_isolation_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(REQUIRED_COLUMNS)
            writer.writerow(["CHATGPT", "AI", "ChatGPT", "private", "30D", 100, 7, "success1@example.com|pass", "", 1])
            writer.writerow(["VEO3", "AI", "VEO3", "private", "30D", 100, 7, "missing@example.com|pass", "", 1])
            writer.writerow(["CHATGPT", "AI", "ChatGPT", "private", "30D", 100, 7, "success2@example.com|pass", "", 1])
        row_isolation = import_inventory(row_isolation_path, self.db_path)
        self.assertEqual(
            (row_isolation["products_created"], row_isolation["credentials_added"], row_isolation["invalid_rows"]),
            (0, 2, 1),
        )
        with closing(sqlite3.connect(self.db_path)) as connection:
            self.assertIsNone(connection.execute("SELECT id FROM products WHERE code = 'VEO3'").fetchone())
        self.assertEqual(self.store.get_stock_count("CHATGPT"), 2)

    def test_import_workbook_multi_sheet_reports_canonical_codes(self) -> None:
        workbook_path = Path(self.temp_dir.name) / "multi-sheet.xlsx"
        workbook = Workbook()
        default_sheet = workbook.active
        default_sheet.title = "README"
        default_sheet.append(["ignored"])
        for sheet_name, credentials in {
            "GEMINI": ("gemini1@example.com|pass", "gemini2@example.com|pass|2fa"),
            "CAPCUT": ("capcut1@example.com|pass", "capcut2@example.com|pass"),
        }.items():
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(REQUIRED_COLUMNS)
            for credential in credentials:
                sheet.append([sheet_name, "AI", sheet_name, "personal", "30D", 100, 7, credential, "", 1])
        workbook.save(workbook_path)
        workbook.close()

        report = import_inventory(workbook_path, self.db_path)

        self.assertEqual((report["credentials_added"], report["row_errors"]), (4, 0))
        self.assertEqual(report["stock"], {"CAPCUT": 2, "GEMINI": 2})
        with closing(sqlite3.connect(self.db_path)) as connection:
            capcut = connection.execute("SELECT code, name, category, delivery_type, price_vnd FROM products WHERE code = 'CAPCUT'").fetchone()
        self.assertEqual(capcut, ("CAPCUT", "CAPCUT PRO", "account", "account", 400000))

    def test_replace_preserves_reserved_and_delivered_inventory(self) -> None:
        workbook_path = Path(self.temp_dir.name) / "replace.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "GEMINI"
        sheet.append(REQUIRED_COLUMNS)
        sheet.append(["GEMINI", "AI", "Gemini", "personal", "30D", 100, 7, "new1@example.com|pass", "", 1])
        sheet.append(["GEMINI", "AI", "Gemini", "personal", "30D", 100, 7, "new2@example.com|pass", "", 1])
        workbook.save(workbook_path)
        workbook.close()

        available_id = self.store.add_inventory_item("GEMINI", "old-available@example.com|pass")
        reserved_id = self.store.add_inventory_item("GEMINI", "old-reserved@example.com|pass")
        delivered_id = self.store.add_inventory_item("GEMINI", "old-delivered@example.com|pass")
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                connection.execute("UPDATE inventory_items SET status = 'reserved' WHERE id = ?", (reserved_id,))
                connection.execute("UPDATE inventory_items SET status = 'delivered' WHERE id = ?", (delivered_id,))

        report = import_inventory(workbook_path, self.db_path, mode="replace")

        self.assertEqual((report["credentials_added"], report["row_errors"]), (2, 0))
        with closing(sqlite3.connect(self.db_path)) as connection:
            states = dict(connection.execute(
                "SELECT id, status FROM inventory_items WHERE id IN (?, ?, ?)",
                (available_id, reserved_id, delivered_id),
            ).fetchall())
        self.assertEqual(states[available_id], "disabled")
        self.assertEqual(states[reserved_id], "reserved")
        self.assertEqual(states[delivered_id], "delivered")

    def test_capcut_auto_created_when_missing_product(self) -> None:
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                connection.execute("DELETE FROM inventory_items WHERE product_id IN (SELECT id FROM products WHERE code = 'CAPCUT')")
                connection.execute("DELETE FROM products WHERE code = 'CAPCUT'")
        workbook_path = Path(self.temp_dir.name) / "capcut.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CAPCUT"
        sheet.append(REQUIRED_COLUMNS)
        sheet.append(["", "AI", "CapCut Pro", "personal", "365D", 400000, 365, "capcut@example.com|pass", "", 1])
        workbook.save(workbook_path)
        workbook.close()

        report = import_inventory(workbook_path, self.db_path)

        self.assertEqual((report["products_created"], report["credentials_added"], report["row_errors"]), (1, 1, 0))
        self.assertEqual(report["stock"], {"CAPCUT": 1})

    def test_replace_reimport_same_workbook_does_not_double_stock(self) -> None:
        workbook_path = Path(self.temp_dir.name) / "repeat-replace.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "CAPCUT"
        sheet.append(REQUIRED_COLUMNS)
        for credential in ("one@example.com|pass", "two@example.com|pass", "three@example.com|pass"):
            sheet.append(["CAPCUT", "AI", "CAPCUT", "personal", "365D", 400000, 365, credential, "", 1])
        workbook.save(workbook_path)
        workbook.close()

        first = import_inventory(workbook_path, self.db_path, mode="replace")
        second = import_inventory(workbook_path, self.db_path, mode="replace")

        self.assertEqual((first["credentials_added"], first["row_errors"]), (3, 0))
        self.assertEqual((second["credentials_added"], second["row_errors"]), (3, 0))
        self.assertEqual(second["stock"], {"CAPCUT": 3})

    def test_telegram_catalog_menu_and_order_use_sqlite_price_and_stock(self) -> None:
        for code in ("GEMINI", "GEM-AIPRO-1M-PRIVATE", "CAPCUT", "CAPCUT-PRO-1M-PRIVATE"):
            with closing(sqlite3.connect(self.db_path)) as connection:
                with connection:
                    connection.execute("DELETE FROM inventory_items WHERE product_id IN (SELECT id FROM products WHERE code = ?)", (code,))
                    connection.execute("DELETE FROM products WHERE code = ?", (code,))

        workbook_path = Path(self.temp_dir.name) / "telegram-catalog.xlsx"
        workbook = Workbook()
        gemini = workbook.active
        gemini.title = "GEMINI"
        gemini.append(REQUIRED_COLUMNS)
        for index in range(8):
            gemini.append(["GEMINI", "AI", "Gemini AI Pro", "personal", "30D", 70000, 30, f"gemini{index}@example.com|pass", "", 1])
        capcut = workbook.create_sheet("CAPCUT")
        capcut.append(REQUIRED_COLUMNS)
        for index in range(3):
            capcut.append(["CAPCUT", "AI", "CAPCUT PRO", "personal", "365D", 400000, 365, f"capcut{index}@example.com|pass", "", 1])
        workbook.save(workbook_path)
        workbook.close()

        report = import_inventory(workbook_path, self.db_path, mode="replace")
        self.assertEqual(report["stock"], {"GEMINI": 8, "CAPCUT": 3})
        with closing(sqlite3.connect(self.db_path)) as connection:
            now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
            with connection:
                connection.execute(
                    """
                    INSERT INTO products
                        (id, code, name, active, delivery_type, created_at, updated_at,
                         category, category_key, price_vnd, product_group)
                    VALUES (?, ?, ?, 1, 'account', ?, ?, 'AI', 'AI', 0, 'account')
                    """,
                    ("old-ai-category", "OLD-AI-CATEGORY", "AI", now, now),
                )
                connection.execute(
                    """
                    INSERT INTO products
                        (id, code, name, active, delivery_type, created_at, updated_at,
                         category_key, price_vnd, product_group)
                    VALUES (?, ?, ?, 1, 'account', ?, ?, 'CAPCUT', 0, 'account')
                    """,
                    ("catalog-capcut-zero", "CATALOG-CAPCUT-ZERO", "CAPCUT PRO", now, now),
                )
                connection.execute("UPDATE products SET category_key = 'AI' WHERE code = 'GEMINI'")
                connection.execute("UPDATE products SET name = 'CAPCUT' WHERE code = 'CAPCUT'")
            StoreRepository(self.db_path)
            prices = dict(connection.execute("SELECT code, price_vnd FROM products WHERE code IN ('GEMINI', 'CAPCUT')").fetchall())
            names = dict(connection.execute("SELECT code, name FROM products WHERE code IN ('GEMINI', 'CAPCUT')").fetchall())
        self.assertEqual(prices["GEMINI"], 70000)
        self.assertEqual(names["CAPCUT"], "CAPCUT PRO")

        previous_store_db_path = os.environ.get("STORE_DB_PATH")
        previous_make_order_id = bot._make_order_id
        try:
            os.environ["STORE_DB_PATH"] = str(self.db_path)
            menu_buttons = [button.text for row in bot._product_menu_keyboard().inline_keyboard for button in row]
            self.assertIn("🟢 GEMINI AI (8)", menu_buttons)
            self.assertIn("🟢 CAPCUT PRO (3)", menu_buttons)
            self.assertNotIn("🟢 AI (8)", menu_buttons)
            self.assertFalse(any(label in {"🟢 AI (0)", "🔴 AI (0)", "🟢 AI (8)", "🔴 AI (8)"} for label in menu_buttons))
            self.assertFalse(any("90.000" in label or "199.000" in label or "499.000" in label for label in menu_buttons))

            package_buttons = [button.text for row in bot._package_keyboard("GEMINI AI").inline_keyboard for button in row]
            self.assertIn("🎁 Gemini AI Pro\n💰 70.000đ\n📦 Còn: 8", package_buttons)
            self.assertFalse(any("90.000" in label or "199.000" in label or "499.000" in label for label in package_buttons))
            capcut_package_buttons = [button.text for row in bot._package_keyboard("CAPCUT PRO").inline_keyboard for button in row]
            self.assertIn("🎁 CAPCUT PRO\n💰 400.000đ\n📦 Còn: 3", capcut_package_buttons)
            self.assertNotIn("🎁 CAPCUT PRO\n💰 0đ\n📦 Còn: 0", capcut_package_buttons)

            package = bot._get_package_info("GEMINI AI", "GEMINI")
            self.assertIsNotNone(package)
            self.assertEqual((package["price_vnd"], package["available_count"]), (70000, 8))
            self.assertEqual(bot._menu_available_count("GEMINI AI"), 8)
            self.assertEqual(bot._menu_available_count("CAPCUT PRO"), 3)

            bot._make_order_id = lambda _product_name: "ORD-GEMINI-1"
            fake_update = type(
                "FakeOrderUpdate",
                (),
                {"effective_user": type("FakeUser", (), {"id": 42, "full_name": "Test User", "username": ""})()},
            )()
            order = bot._create_sales_order(fake_update, "GEMINI AI", "GEMINI", 1)
            self.assertEqual(order["unit_price"], 70000)
            self.assertEqual(order["total"], 70000)
            persisted = StoreRepository(self.db_path).find_order("ORD-GEMINI-1")
            self.assertIsNotNone(persisted)
            self.assertEqual((persisted["unit_price_vnd"], persisted["total_vnd"], persisted["quantity"]), (70000, 70000, 1))
            self.assertEqual(StoreRepository(self.db_path).get_stock_count("GEMINI"), 7)
        finally:
            bot._make_order_id = previous_make_order_id
            if previous_store_db_path is None:
                os.environ.pop("STORE_DB_PATH", None)
            else:
                os.environ["STORE_DB_PATH"] = previous_store_db_path

    def test_quantity_reservation_boundaries_follow_available_stock(self) -> None:
        def fresh_repo_with_stock(stock: int) -> StoreRepository:
            db_path = Path(self.temp_dir.name) / f"gemini-stock-{stock}.db"
            with closing(sqlite3.connect(PROJECT_ROOT / "database" / "store.db")) as source, closing(sqlite3.connect(db_path)) as target:
                source.backup(target)
            repo = StoreRepository(db_path)
            with closing(sqlite3.connect(db_path)) as connection:
                with connection:
                    product = connection.execute(
                        "SELECT id FROM products WHERE code IN ('GEMINI', 'GEM-AIPRO-1M-PRIVATE') ORDER BY CASE WHEN code = 'GEMINI' THEN 0 ELSE 1 END LIMIT 1"
                    ).fetchone()
                    self.assertIsNotNone(product)
                    connection.execute("DELETE FROM inventory_items WHERE product_id = ?", (product[0],))
            for index in range(stock):
                repo.add_inventory_item("GEMINI", f"qty{stock}-{index}@example.com|pass")
            self.assertEqual(repo.get_stock_count("GEMINI"), stock)
            return repo

        now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        cases = ((1, 8, True), (5, 3, False), (8, 8, True), (9, 8, False))
        for requested, stock, should_succeed in cases:
            repo = fresh_repo_with_stock(stock)
            order_id = f"ORD-QTY-{stock}-{requested}"
            if should_succeed:
                reserved = repo.create_pending_account_order_and_reserve(
                    order_id=order_id,
                    telegram_user_id=1,
                    username="Quantity Test",
                    product_code="GEMINI",
                    product_name="GEMINI AI",
                    package_name="Gemini AI Pro",
                    quantity=requested,
                    unit_price_vnd=70000,
                    total_vnd=70000 * requested,
                    created_at=now,
                    expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
                )
                self.assertEqual(len(reserved), requested)
                self.assertEqual(repo.get_stock_count("GEMINI"), stock - requested)
            else:
                with self.assertRaisesRegex(ValueError, "Insufficient available inventory"):
                    repo.create_pending_account_order_and_reserve(
                        order_id=order_id,
                        telegram_user_id=1,
                        username="Quantity Test",
                        product_code="GEMINI",
                        product_name="GEMINI AI",
                        package_name="Gemini AI Pro",
                        quantity=requested,
                        unit_price_vnd=70000,
                        total_vnd=70000 * requested,
                        created_at=now,
                        expire_at=(datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
                    )

    def test_catalog_schema_and_optional_import_columns(self) -> None:
        with closing(sqlite3.connect(self.db_path)) as connection:
            columns = {row[1] for row in connection.execute("PRAGMA table_info(products)")}
        self.assertTrue({"menu_order", "show_in_menu", "product_group", "category_key", "description"}.issubset(columns))

        catalog_path = Path(self.temp_dir.name) / "catalog.csv"
        with catalog_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["product_code", "email", "password", "2fa"])
            writer.writerow(["CHATGPT", "a@example.com", "pass", "2fa"])
            writer.writerow(["NOT_ALLOWED", "b@example.com", "pass", "2fa"])
        report = import_inventory(catalog_path, self.db_path)
        self.assertEqual((report["credentials_added"], report["invalid_rows"]), (1, 1))
        self.assertTrue(any("Unsupported product_code: NOT_ALLOWED" in error for error in report["errors"]))

        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                for code, name, category_key in (
                    ("CATEGORY-CHATGPT", "ChatGPT Plus", "CHATGPT"),
                    ("CATEGORY-GEMINI", "Gemini Advanced", "GEMINI"),
                    ("CATEGORY-GROK", "Grok Super", "GROK"),
                ):
                    connection.execute(
                        """
                        INSERT INTO products
                            (id, code, name, category, category_key, active, delivery_type, created_at, updated_at)
                        VALUES (?, ?, ?, 'AI', ?, 1, 'account', ?, ?)
                        """,
                        (f"test-{code}", code, name, category_key, datetime.now(timezone.utc).isoformat(), datetime.now(timezone.utc).isoformat()),
                    )
        categories = {row["category_key"] for row in self.store.list_visible_categories()}
        self.assertTrue({"CHATGPT", "GEMINI", "GROK"}.issubset(categories))
        self.assertNotIn("AI", categories)
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                now = datetime.now(timezone.utc).isoformat()
                connection.execute(
                    "INSERT INTO products (id, code, name, category_key, active, delivery_type, created_at, updated_at) VALUES (?, ?, ?, ?, 1, 'account', ?, ?)",
                    ("mixed-case-chatgpt", "MIXED-CHATGPT", "ChatGPT Alternate", "ChatGPT", now, now),
                )
        self.assertEqual(
            [row["category_key"] for row in self.store.list_visible_categories()].count("CHATGPT"),
            1,
        )

    def test_cleanup_demo_inventory_preserves_real_reserved_and_delivered(self) -> None:
        now = datetime.now(timezone.utc).isoformat()
        with closing(sqlite3.connect(self.db_path)) as connection:
            with connection:
                for product_id, code, note in (("demo-product", "DEMO-CODE", "Example only"), ("real-product", "REAL-CODE", "real supplier stock"), ("protected-demo-code", "DEMO-PROTECTED", "real supplier stock")):
                    connection.execute(
                        "INSERT INTO products (id, code, name, note, active, delivery_type, created_at, updated_at) VALUES (?, ?, ?, ?, 1, 'account', ?, ?)",
                        (product_id, code, code, note, now, now),
                    )
                for item_id, product_id, status in (("demo-available", "demo-product", "available"), ("demo-delivered", "demo-product", "delivered"), ("real-available", "real-product", "available"), ("protected-available", "protected-demo-code", "available")):
                    connection.execute(
                        "INSERT INTO inventory_items (id, product_id, secret_value, status, created_at) VALUES (?, ?, ?, ?, ?)",
                        (item_id, product_id, item_id + "|pass", status, now),
                    )
        self.assertEqual(cleanup_demo_inventory(self.db_path, {"DEMO-CODE", "DEMO-PROTECTED"}), 1)
        with closing(sqlite3.connect(self.db_path)) as connection:
            states = dict(connection.execute("SELECT id, status FROM inventory_items WHERE id IN ('demo-available', 'demo-delivered', 'real-available', 'protected-available')").fetchall())
        self.assertEqual(states, {"demo-available": "disabled", "demo-delivered": "delivered", "real-available": "available", "protected-available": "available"})
        stock = {row["product_code"]: row for row in self.store.get_stock_summary()}
        self.assertEqual(stock["DEMO-CODE"]["available"], 0)


if __name__ == "__main__":
    unittest.main(verbosity=2)
