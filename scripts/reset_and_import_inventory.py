from __future__ import annotations

import argparse
import sqlite3
import sys
import uuid
from collections import Counter, defaultdict
from contextlib import closing
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from repository.store_repository import ACCOUNT_PRODUCT_CODE_ALIASES, canonical_catalog_product
from scripts.import_inventory import (
    IMPORT_COLUMNS,
    SKIPPED_SHEET_NAMES,
    canonical_price_vnd,
    credential_from_row,
    ensure_product,
    parse_int,
    product_display_name,
    sync_product_metadata,
    text,
)


TARGET_PRODUCTS = ("CHATGPT", "GEMINI", "CAPCUT")
RESET_STATUSES = ("available", "reserved", "disabled")
GEMINI_MAX_SLOTS_PER_CREDENTIAL = 2


@dataclass(frozen=True)
class InventoryRow:
    product_code: str
    credential: str
    row: dict[str, Any]
    source: str
    row_number: int
    slot: int = 1


@dataclass(frozen=True)
class PreparedInventory:
    rows_by_product: dict[str, list[InventoryRow]]
    summary: dict[str, dict[str, int]]
    warnings: list[str]


@dataclass(frozen=True)
class ImportRowResult:
    item: InventoryRow
    identity: str
    status: str
    reason: str = ""
    inventory_item_id: str = ""


@dataclass(frozen=True)
class ExcelReadResult:
    rows: list[InventoryRow]
    invalid_rows: list[ImportRowResult]


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def normalize_product_code(value: Any) -> str:
    normalized = " ".join(text(value).upper().split())
    compact = normalized.replace(" ", "")
    if compact == "CHATGPT":
        return "CHATGPT"
    if compact in {"GEMINI", "CAPCUT"}:
        return compact
    if normalized in {"CAPCUT PRO", "GEMINI AI", "CHATGPT PLUS"}:
        return {"CAPCUT PRO": "CAPCUT", "GEMINI AI": "GEMINI", "CHATGPT PLUS": "CHATGPT"}[normalized]
    for canonical, aliases in ACCOUNT_PRODUCT_CODE_ALIASES.items():
        if normalized == canonical or compact == canonical.replace(" ", ""):
            return normalize_product_code(canonical)
        if normalized in {alias.upper() for alias in aliases}:
            return normalize_product_code(canonical)
    return compact or normalized


def product_code_candidates(product_code: str) -> list[str]:
    candidates = [product_code]
    for alias in ACCOUNT_PRODUCT_CODE_ALIASES.get(product_code, ()):
        alias = str(alias).strip().upper()
        if alias and alias not in candidates:
            candidates.append(alias)
    return candidates


def raw_identity_from_row(row: dict[str, Any]) -> str:
    credential_text = text(row.get("credential_text"))
    if credential_text:
        return credential_text
    parts = [text(row.get("email")), text(row.get("password"))]
    if text(row.get("2fa")) or text(row.get("recovery_email")):
        parts.append(text(row.get("2fa")))
    if text(row.get("recovery_email")):
        parts.append(text(row.get("recovery_email")))
    return "|".join(parts)


def read_excel_rows(path: Path) -> ExcelReadResult:
    if not path.is_file():
        raise FileNotFoundError(f"Excel file not found: {path}")
    if path.suffix.lower() != ".xlsx":
        raise ValueError(f"Input must be an .xlsx file: {path}")

    rows: list[InventoryRow] = []
    invalid_rows: list[ImportRowResult] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            sheet_code = normalize_product_code(sheet.title)
            if text(sheet.title).upper() in SKIPPED_SHEET_NAMES:
                continue
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue
            header = [text(value).lower() for value in header_row]
            positions = {name: index for index, name in enumerate(header) if name}
            sheet_is_target_product = sheet_code in TARGET_PRODUCTS
            if not sheet_is_target_product and "product_code" not in positions:
                continue
            if "credential_text" not in positions and not {"email", "password"}.issubset(positions):
                raise ValueError(f"{path.name} / {sheet.title}: missing credential columns")

            for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(value is not None and text(value) for value in values):
                    continue
                row = {
                    column: values[positions[column]] if column in positions and positions[column] < len(values) else ""
                    for column in IMPORT_COLUMNS
                }
                raw_product_code = sheet_code if sheet_is_target_product else row.get("product_code")
                product_code = normalize_product_code(raw_product_code)
                if product_code not in TARGET_PRODUCTS:
                    continue
                row["product_code"] = product_code
                row["category_key"] = product_code
                row["product_group"] = "account"
                row["category"] = "account"
                base_item = InventoryRow(
                    product_code=product_code,
                    credential=raw_identity_from_row(row),
                    row=row,
                    source=path.name,
                    row_number=row_number,
                )
                try:
                    credential = credential_from_row(row)
                except ValueError as exc:
                    invalid_rows.append(
                        ImportRowResult(
                            item=base_item,
                            identity=base_item.credential,
                            status="invalid skipped",
                            reason=str(exc),
                        )
                    )
                    continue
                rows.append(
                    InventoryRow(
                        product_code=product_code,
                        credential=credential,
                        row=row,
                        source=path.name,
                        row_number=row_number,
                    )
                )
    finally:
        workbook.close()
    return ExcelReadResult(rows=rows, invalid_rows=invalid_rows)


def gemini_slot_credential(credential: str, slot: int) -> str:
    return f"{credential}|GEMINI_SLOT_{slot}"


def prepare_inventory(rows: list[InventoryRow]) -> PreparedInventory:
    rows_by_product: dict[str, list[InventoryRow]] = {code: [] for code in TARGET_PRODUCTS}
    summary: dict[str, dict[str, int]] = {
        code: {"rows": 0, "unique": 0, "sellable": 0, "duplicate": 0, "extra_skipped": 0}
        for code in TARGET_PRODUCTS
    }
    warnings: list[str] = []
    seen_standard: dict[str, set[str]] = {code: set() for code in TARGET_PRODUCTS}
    gemini_seen_count: Counter[str] = Counter()

    for item in rows:
        summary[item.product_code]["rows"] += 1
        if item.product_code == "GEMINI":
            if gemini_seen_count[item.credential] == 0:
                summary["GEMINI"]["unique"] += 1
            gemini_seen_count[item.credential] += 1
            slot = gemini_seen_count[item.credential]
            if slot > GEMINI_MAX_SLOTS_PER_CREDENTIAL:
                summary["GEMINI"]["extra_skipped"] += 1
                warnings.append(
                    f"GEMINI extra slot skipped: credential at {item.source}:{item.row_number} exceeds "
                    f"{GEMINI_MAX_SLOTS_PER_CREDENTIAL} slots"
                )
                continue
            item_row = dict(item.row)
            item_row["note"] = f"{text(item_row.get('note'))} GEMINI_SLOT_{slot}".strip()
            rows_by_product["GEMINI"].append(
                InventoryRow(
                    product_code=item.product_code,
                    credential=item.credential,
                    row=item_row,
                    source=item.source,
                    row_number=item.row_number,
                    slot=slot,
                )
            )
            summary["GEMINI"]["sellable"] += 1
        elif item.credential in seen_standard[item.product_code]:
            summary[item.product_code]["duplicate"] += 1
        else:
            seen_standard[item.product_code].add(item.credential)
            summary[item.product_code]["unique"] += 1
            summary[item.product_code]["sellable"] += 1
            rows_by_product[item.product_code].append(item)
    return PreparedInventory(rows_by_product=rows_by_product, summary=summary, warnings=warnings)


def debug_chatgpt_excel_rows(rows: list[InventoryRow], invalid_rows: list[ImportRowResult]) -> None:
    print()
    print("DEBUG CHATGPT EXCEL ROWS:")
    chatgpt_items: list[tuple[int, InventoryRow, ImportRowResult | None]] = [
        (item.row_number, item, None)
        for item in rows
        if item.product_code == "CHATGPT"
    ]
    chatgpt_items.extend(
        (result.item.row_number, result.item, result)
        for result in invalid_rows
        if result.item.product_code == "CHATGPT"
    )
    for _, item, invalid in sorted(chatgpt_items, key=lambda value: value[0]):
        credential_parts = item.credential.split("|")
        password_present = len(credential_parts) >= 2 and bool(text(credential_parts[1]))
        secret_present = len(credential_parts) >= 3 and bool(text(credential_parts[2]))
        status = "invalid" if invalid else "valid"
        reason = f" reason={invalid.reason}" if invalid else ""
        print(
            f"row={item.row_number} product_code={item.product_code} "
            f"identity={item.credential} password_present={password_present} "
            f"secret_2fa_present={secret_present} normalized_identity={item.credential} "
            f"status={status}{reason}"
        )


def find_product_by_code(connection: sqlite3.Connection, product_code: str) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT id, code FROM products WHERE code = ?",
        (product_code,),
    ).fetchone()


def reset_target_inventory(connection: sqlite3.Connection) -> tuple[int, int]:
    target_codes: list[str] = []
    for product_code in TARGET_PRODUCTS:
        for candidate in product_code_candidates(product_code):
            if candidate not in target_codes:
                target_codes.append(candidate)

    placeholders = ",".join("?" for _ in target_codes)
    status_placeholders = ",".join("?" for _ in RESET_STATUSES)
    item_rows = connection.execute(
        f"""
        SELECT i.id
        FROM inventory_items AS i
        JOIN products AS p ON p.id = i.product_id
        WHERE UPPER(p.code) IN ({placeholders})
          AND i.status IN ({status_placeholders})
        """,
        (*target_codes, *RESET_STATUSES),
    ).fetchall()
    item_ids = [str(row["id"]) for row in item_rows]
    if not item_ids:
        return 0, 0

    item_placeholders = ",".join("?" for _ in item_ids)
    link_count = connection.execute(
        f"SELECT COUNT(*) AS count FROM order_inventory_items WHERE inventory_item_id IN ({item_placeholders})",
        item_ids,
    ).fetchone()["count"]
    connection.execute(
        f"DELETE FROM order_inventory_items WHERE inventory_item_id IN ({item_placeholders})",
        item_ids,
    )
    connection.execute(
        f"DELETE FROM inventory_movements WHERE inventory_item_id IN ({item_placeholders})",
        item_ids,
    )
    connection.execute(
        f"DELETE FROM inventory_items WHERE id IN ({item_placeholders})",
        item_ids,
    )
    return int(link_count or 0), len(item_ids)


def ensure_target_product(connection: sqlite3.Connection, product_code: str, row: dict[str, Any], now: str) -> sqlite3.Row:
    product = find_product_by_code(connection, product_code)
    if product:
        sync_product_metadata(connection, product["id"], product_code, row, now)
        return product
    return ensure_product(connection, product_code, row, now)


def inventory_secret_value(item: InventoryRow) -> str:
    if item.product_code == "GEMINI":
        return gemini_slot_credential(item.credential, item.slot)
    return item.credential


def import_prepared_rows(
    connection: sqlite3.Connection,
    rows_by_product: dict[str, list[InventoryRow]],
    invalid_rows: list[ImportRowResult],
) -> tuple[dict[str, int], list[ImportRowResult]]:
    imported = {code: 0 for code in TARGET_PRODUCTS}
    row_results: list[ImportRowResult] = list(invalid_rows)
    inserted_identities: dict[str, set[str]] = {code: set() for code in TARGET_PRODUCTS}
    now = utc_now_iso()
    for product_code in TARGET_PRODUCTS:
        product_rows = rows_by_product[product_code]
        if not product_rows:
            continue
        product = ensure_target_product(connection, product_code, product_rows[0].row, now)
        for item in product_rows:
            identity = inventory_secret_value(item)
            if identity in inserted_identities[product_code]:
                row_results.append(
                    ImportRowResult(
                        item=item,
                        identity=identity,
                        status="duplicate skipped",
                        reason="duplicate identity in this import batch",
                    )
                )
                continue
            item_id = str(uuid.uuid4())
            connection.execute(
                "INSERT INTO inventory_items (id, product_id, secret_value, status, created_at) VALUES (?, ?, ?, 'available', ?)",
                (item_id, product["id"], identity, utc_now_iso()),
            )
            connection.execute(
                """
                INSERT INTO inventory_movements
                    (id, inventory_item_id, action, source, created_at)
                VALUES (?, ?, 'import', 'reset_and_import_inventory', ?)
                """,
                (str(uuid.uuid4()), item_id, utc_now_iso()),
            )
            imported[product_code] += 1
            inserted_identities[product_code].add(identity)
            row_results.append(
                ImportRowResult(
                    item=item,
                    identity=identity,
                    status="inserted",
                    inventory_item_id=item_id,
                )
            )
    return imported, row_results


def actual_available(connection: sqlite3.Connection) -> dict[str, int]:
    actual = {code: 0 for code in TARGET_PRODUCTS}
    for product_code in TARGET_PRODUCTS:
        row = connection.execute(
            """
            SELECT COUNT(*) AS count
            FROM inventory_items AS i
            JOIN products AS p ON p.id = i.product_id
            WHERE UPPER(p.code) = ? AND i.status = 'available'
            """,
            (product_code,),
        ).fetchone()
        actual[product_code] = int(row["count"] or 0)
    return actual


def print_chatgpt_import_results(row_results: list[ImportRowResult]) -> None:
    print()
    print("DEBUG CHATGPT IMPORT RESULTS:")
    for result in sorted(
        [item for item in row_results if item.item.product_code == "CHATGPT"],
        key=lambda value: value.item.row_number,
    ):
        detail = f" reason={result.reason}" if result.reason else ""
        item_id = f" inventory_item_id={result.inventory_item_id}" if result.inventory_item_id else ""
        print(
            f"row={result.item.row_number} identity={result.identity} "
            f"result={result.status}{item_id}{detail}"
        )


def missing_available_rows(connection: sqlite3.Connection, row_results: list[ImportRowResult]) -> list[ImportRowResult]:
    missing: list[ImportRowResult] = []
    for result in row_results:
        if result.item.product_code != "CHATGPT":
            continue
        if result.status != "inserted":
            missing.append(result)
            continue
        row = connection.execute(
            """
            SELECT COUNT(*) AS count
            FROM inventory_items AS i
            JOIN products AS p ON p.id = i.product_id
            WHERE UPPER(p.code) = 'CHATGPT'
              AND i.status = 'available'
              AND i.secret_value = ?
            """,
            (result.identity,),
        ).fetchone()
        if int(row["count"] or 0) <= 0:
            missing.append(
                ImportRowResult(
                    item=result.item,
                    identity=result.identity,
                    status="not available after insert",
                    reason="no matching CHATGPT available inventory_item found",
                    inventory_item_id=result.inventory_item_id,
                )
            )
    return missing


def print_missing_available_rows(missing: list[ImportRowResult]) -> None:
    if not missing:
        return
    print()
    print("DEBUG CHATGPT MISSING AVAILABLE ROWS:")
    for result in sorted(missing, key=lambda value: value.item.row_number):
        detail = f" reason={result.reason}" if result.reason else ""
        print(
            f"row={result.item.row_number} source={result.item.source} "
            f"identity={result.identity} status={result.status}{detail}"
        )


def print_expected(summary: dict[str, dict[str, int]]) -> None:
    print("EXPECTED FROM EXCEL:")
    for code in TARGET_PRODUCTS:
        print(f"{code} rows: {summary[code]['rows']}")
        print(f"{code} unique credentials/email: {summary[code]['unique']}")
        print(f"{code} sellable items: {summary[code]['sellable']}")


def print_report(
    summary: dict[str, dict[str, int]],
    imported: dict[str, int],
    actual: dict[str, int],
    deleted_links: int,
    deleted_items: int,
) -> None:
    print()
    print("RESET RESULT:")
    print(f"Deleted inventory links: {deleted_links}")
    print(f"Deleted inventory items: {deleted_items}")
    print()
    print("IMPORT RESULT:")
    for code in TARGET_PRODUCTS:
        print(f"{code} rows in Excel: {summary[code]['rows']}")
        print(f"{code} unique credentials/email: {summary[code]['unique']}")
        print(f"{code} sellable items: {summary[code]['sellable']}")
        print(f"{code} imported: {imported[code]}")
        if code == "GEMINI":
            print(f"{code} extra slots skipped: {summary[code]['extra_skipped']}")
        else:
            print(f"{code} duplicate skipped: {summary[code]['duplicate']}")
        print()
    print("ACTUAL IN DATABASE:")
    for code in TARGET_PRODUCTS:
        print(f"{code} available: {actual[code]}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Reset and import CHATGPT, GEMINI, and CAPCUT inventory from Excel.")
    parser.add_argument("--database", required=True, type=Path)
    parser.add_argument("--chatgpt-file", required=True, type=Path)
    parser.add_argument("--multi-file", required=True, type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    chatgpt_read = read_excel_rows(args.chatgpt_file)
    multi_read = read_excel_rows(args.multi_file)
    rows = chatgpt_read.rows + multi_read.rows
    invalid_rows = chatgpt_read.invalid_rows + multi_read.invalid_rows
    debug_chatgpt_excel_rows(rows, invalid_rows)
    prepared = prepare_inventory(rows)
    print_expected(prepared.summary)
    for warning in prepared.warnings:
        print(f"WARNING: {warning}")

    if not args.database.is_file():
        raise FileNotFoundError(f"Store database not found: {args.database}")

    with closing(sqlite3.connect(args.database)) as connection:
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        try:
            connection.execute("BEGIN IMMEDIATE")
            deleted_links, deleted_items = reset_target_inventory(connection)
            imported, row_results = import_prepared_rows(connection, prepared.rows_by_product, invalid_rows)
            print_chatgpt_import_results(row_results)
            actual = actual_available(connection)
            missing_chatgpt = missing_available_rows(connection, row_results)
            print_missing_available_rows(missing_chatgpt)
            failures = [
                f"{code} available={actual[code]} expected={prepared.summary[code]['sellable']}"
                for code in TARGET_PRODUCTS
                if actual[code] != prepared.summary[code]["sellable"]
            ]
            if missing_chatgpt:
                failures.append(
                    "CHATGPT rows not available: "
                    + ", ".join(f"row {item.item.row_number} ({item.status})" for item in missing_chatgpt)
                )
            if failures:
                raise RuntimeError("Validation failed: " + "; ".join(failures))
            connection.commit()
        except Exception:
            connection.rollback()
            raise

    print_report(prepared.summary, imported, actual, deleted_links, deleted_items)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
